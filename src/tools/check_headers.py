# ==========================================================
# DataPulse – check_headers.py
# Detector quirúrgico de encabezados por hoja
# ==========================================================
import sys
from pathlib import Path

# === FIX DE RUTA GLOBAL ===
BASE_DIR = Path(__file__).resolve().parents[3]
sys.path.append(str(BASE_DIR))

import json
import pandas as pd
from openpyxl import load_workbook
from src.core.config import EXCELS
from src.core.logger import log

# === CONFIGURACIÓN ===
EXCEL_PATH = Path(EXCELS.get("movimientos", ""))
OUTPUT_JSON = Path(__file__).resolve().parent / "headers_detected.json"
OUTPUT_XLSX = Path(__file__).resolve().parent / "headers_detected.xlsx"
OMITIR_HOJAS = {"fe", "reporte_bancos", "reporte_banco", "reportes_bancos"}  # omite variantes
HEADER_ROW = 3  # Fila 4 visible en Excel

# ==========================================================
def limpiar_columna(nombre):
    """Limpieza mínima sin alterar el formato visible."""
    return str(nombre).replace("\n", " ").replace("  ", " ").strip()

# ==========================================================
def detectar_encabezados():
    """Detecta y exporta los encabezados de cada hoja."""
    if not EXCEL_PATH.exists():
        log(f"❌ No se encontró el archivo Excel: {EXCEL_PATH}")
        return

    log(f"📘 Analizando encabezados del archivo: {EXCEL_PATH.name}")

    wb = load_workbook(EXCEL_PATH, read_only=True)
    hojas = [h for h in wb.sheetnames if h.strip().lower() not in OMITIR_HOJAS]
    wb.close()

    resultados = {}
    registros_para_excel = []

    for hoja in hojas:
        try:
            df = pd.read_excel(
                EXCEL_PATH,
                sheet_name=hoja,
                header=HEADER_ROW,
                nrows=0,
                engine="openpyxl"
            )
            columnas = [limpiar_columna(c) for c in df.columns]
            resultados[hoja] = columnas

            registros_para_excel.append({
                "Hoja": hoja,
                "CantidadColumnas": len(columnas),
                "Encabezados": ", ".join(columnas)
            })

            log(f"✅ {hoja}: {len(columnas)} columnas detectadas.")
        except Exception as e:
            log(f"⚠️ Error leyendo '{hoja}': {e}")

    # === Guardar JSON ===
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(resultados, f, ensure_ascii=False, indent=4)
    log(f"💾 Archivo JSON generado: {OUTPUT_JSON}")

    # === Guardar Excel ===
    df_export = pd.DataFrame(registros_para_excel)
    df_export.to_excel(OUTPUT_XLSX, index=False)
    log(f"💾 Archivo Excel generado: {OUTPUT_XLSX}")

    log("✅ Análisis completado correctamente.")

# ==========================================================
if __name__ == "__main__":
    detectar_encabezados()
