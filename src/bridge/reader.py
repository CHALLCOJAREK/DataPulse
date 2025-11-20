# src/bridge/reader.py
import sys
from pathlib import Path
sys.path.append(str(Path(__file__).resolve().parents[2]))

import re
import pandas as pd
from openpyxl import load_workbook
from core.config import EXCELS
from core.logger import log

# ==========================================================
# CONFIGURACIÓN BASE
# ==========================================================
OMITIR_HOJAS = {
    "Reporte_Bancos", "FE",     # del Excel 1
    "Tablas", "ER Funcion", "ER Naturaleza", "Partidas_Presupuestos"  # del Excel 2
}
  # Hojas que se omiten por defecto
_currency_re = re.compile(r"[^\d\-,.]+")
NUMERIC_HINTS = {"PARCIAL", "MONTO", "ABONO", "RETIRO", "SALDO"}
DATE_HINTS = {"FECHA"}

# ==========================================================
# FUNCIONES AUXILIARES
# ==========================================================
def _fix_mojibake(text):
    """Corrige errores comunes de codificación (Ã³ → ó, Â → '')."""
    if text is None or not isinstance(text, str):
        return text
    if "Ã" in text or "Â" in text:
        try:
            return text.encode("latin1", errors="ignore").decode("utf-8", errors="ignore")
        except Exception:
            return text
    return text

def _to_number(val):
    """Convierte texto a número flotante seguro."""
    if pd.isna(val):
        return None
    s = str(val).strip()
    if not s or s == "-":
        return None
    s = _currency_re.sub("", s)
    if "," in s and "." in s and s.rfind(",") > s.rfind("."):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None

def _to_date(val):
    """Convierte fechas tipo Excel o '1-Ene' a ISO (YYYY-MM-DD)."""
    if pd.isna(val) or str(val).strip() in ("", "-"):
        return None
    val = str(val).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", val):
        return val
    ts = pd.to_datetime(val, dayfirst=True, errors="coerce")
    return ts.date().isoformat() if pd.notna(ts) else None

def _detect_header_row(ws):
    """Detecta la fila de encabezado real (mínimo 3 columnas con texto)."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        values = [cell.value for cell in row if cell.value not in (None, "")]
        if len(values) >= 3:
            return row[0].row
    return 1

def _col_letter(idx):
    """Convierte índice numérico a letra (1→A, 27→AA)."""
    s = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        s = chr(65 + rem) + s
    return s

# ==========================================================
# LECTOR PRINCIPAL MULTIEXCEL
# ==========================================================
def leer_excel_completo():
    """
    Lee todos los Excels detectados en EXCELS.
    - Detecta encabezados automáticamente.
    - Limpia datos y tipifica sin alterar estructura.
    - Ignora hojas definidas en OMITIR_HOJAS.
    """
    if not EXCELS:
        log("⚠️ No se detectaron archivos Excel configurados en el entorno.")
        return {}

    dataframes = {}
    total_hojas = 0

    for key, excel_path in EXCELS.items():
        if not excel_path.exists():
            log(f"⚠️ Archivo no encontrado: {excel_path}")
            continue

        log(f"📘 Leyendo archivo: {excel_path.name}")

        try:
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            all_sheets = wb.sheetnames
            hojas = [h for h in all_sheets if h not in OMITIR_HOJAS]
            omitidas = [h for h in all_sheets if h in OMITIR_HOJAS]
            wb.close()

            log(f"📄 Hojas detectadas: {len(hojas)} válidas / {len(omitidas)} omitidas")

            if omitidas:
                log(f"   🪶 Omitidas: {', '.join(omitidas)}")

            for hoja in hojas:
                try:
                    wb = load_workbook(excel_path, read_only=True, data_only=True)
                    ws = wb[hoja]

                    header_row = _detect_header_row(ws)
                    max_col = ws.max_column
                    max_row = ws.max_row

                    # Determinar última columna con datos reales
                    last_valid_col = 0
                    for col_idx in range(1, max_col + 1):
                        if any(ws.cell(row=r, column=col_idx).value not in (None, "")
                               for r in range(header_row, min(max_row, header_row + 10))):
                            last_valid_col = col_idx

                    if last_valid_col == 0:
                        log(f"⚠️ Hoja vacía o sin columnas válidas: {hoja}")
                        wb.close()
                        continue

                    last_col_letter = _col_letter(last_valid_col)
                    wb.close()

                    usecols_exact = f"A:{last_col_letter}"

                    df = pd.read_excel(
                        excel_path,
                        sheet_name=hoja,
                        header=header_row - 1,
                        usecols=usecols_exact,
                        engine="openpyxl",
                        dtype=str
                    )

                    # Limpieza y normalización
                    df = df.applymap(_fix_mojibake)
                    df = df.replace({r"^\s*$": None, r"^\s*-\s*$": None}, regex=True)

                    for col in df.columns:
                        upper = str(col).strip().upper()
                        if upper in DATE_HINTS:
                            df[col] = df[col].apply(_to_date)
                        elif upper in NUMERIC_HINTS:
                            df[col] = df[col].apply(_to_number)

                    log(f"✅ Hoja '{hoja}' cargada ({len(df)} filas, {len(df.columns)} columnas)")
                    total_hojas += 1

                    # Guardar con formato: excelX_nombreHoja
                    dataframes[f"{key}_{hoja}"] = df

                except Exception as e:
                    log(f"⚠️ Error al procesar hoja '{hoja}' ({excel_path.name}): {e}")

        except Exception as e:
            log(f"❌ Error al abrir Excel '{excel_path.name}': {e}")

    log(f"\n📊 Lectura completada: {total_hojas} hojas procesadas correctamente.\n")
    return dataframes

# ==========================================================
# EJECUCIÓN DIRECTA
# ==========================================================
if __name__ == "__main__":
    resultados = leer_excel_completo()
    print("\n📘 RESUMEN FINAL DE LECTURA:")
    for hoja, df in resultados.items():
        print(f" • {hoja}: {len(df)} registros | {len(df.columns)} columnas")
